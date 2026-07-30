"""Re-organize scraped CSV data into clean lead sheets."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from scraper.formatter import normalize_row
from scraper.storage import (
    CSV_COLUMNS,
    DATA_DIR,
    FALLBACK_MASTER,
    PENDING_FILE,
    list_data_dates,
    today_folder,
)

ORGANIZED_CSV = "leads_organized.csv"
ORGANIZED_XLSX = "leads_organized.xlsx"


def _read_rows(csv_path: Path) -> list[dict[str, str]]:
    if not csv_path.exists():
        return []
    try:
        with csv_path.open(encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    except (PermissionError, OSError):
        return []


def _dedupe_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    unique: list[dict[str, str]] = []
    for row in reversed(rows):
        key = row.get("place_id") or f"{row.get('name', '')}|{row.get('phone', '')}|{row.get('address', '')}"
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    unique.reverse()
    return unique


def collect_all_rows(folder: Path) -> list[dict[str, str]]:
    """Gather rows from master, fallback, pending, and per-search CSV files."""
    rows: list[dict[str, str]] = []
    skip_names = {
        "all_results.csv",
        FALLBACK_MASTER,
        ORGANIZED_CSV,
    }

    for name in ("all_results.csv", FALLBACK_MASTER):
        rows.extend(_read_rows(folder / name))

    pending = folder / PENDING_FILE
    if pending.exists():
        try:
            for line in pending.read_text(encoding="utf-8").splitlines():
                if line.strip():
                    rows.append(json.loads(line))
        except (json.JSONDecodeError, OSError):
            pass

    for csv_file in sorted(folder.glob("*.csv")):
        if csv_file.name in skip_names:
            continue
        rows.extend(_read_rows(csv_file))

    return rows


def _safe_write_csv(path: Path, rows: list[dict[str, str]]) -> bool:
    try:
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return True
    except (PermissionError, OSError) as exc:
        print(f"  [SKIP] Could not write {path.name}: {exc}")
        print("         Close the file in Excel/editor and run organize.bat again.")
        return False


def organize_folder(folder: Path, rewrite_master: bool = False) -> tuple[int, Path, Path]:
    rows = collect_all_rows(folder)
    cleaned = _dedupe_rows([normalize_row(r) for r in rows])
    cleaned.sort(key=lambda r: (r.get("state", ""), r.get("city", ""), r.get("name", "")))

    csv_out = folder / ORGANIZED_CSV
    _safe_write_csv(csv_out, cleaned)

    xlsx_out = folder / ORGANIZED_XLSX
    try:
        _write_xlsx(cleaned, xlsx_out)
    except (PermissionError, OSError) as exc:
        print(f"  [SKIP] Could not write {xlsx_out.name}: {exc}")

    if rewrite_master:
        master = folder / "all_results.csv"
        if not _safe_write_csv(master, cleaned):
            _safe_write_csv(folder / FALLBACK_MASTER, cleaned)

    return len(cleaned), csv_out, xlsx_out


def _write_xlsx(rows: list[dict[str, str]], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = CSV_COLUMNS
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    widths = {
        "name": 34, "phone": 16, "website": 28, "email": 24, "address": 38,
        "street": 28, "city": 18, "state": 8, "zip": 10, "rating": 8,
        "review_count": 12, "category": 22, "hours_status": 12, "hours_detail": 22,
        "google_maps_url": 36, "search_keyword": 22, "search_place": 22,
        "search_type": 10, "search_state": 10, "scraped_at": 20, "place_id": 24,
    }
    for col, key in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(key, 14)

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    wb.save(path)


def refresh_daily_exports(folder: Path | None = None) -> None:
    folder = folder or today_folder()
    if folder.exists():
        organize_folder(folder, rewrite_master=False)


def print_organize_report(date_str: str | None = None) -> None:
    dates = [date_str] if date_str else list_data_dates()
    if not dates:
        print("No data folders found.")
        return

    for d in dates:
        folder = DATA_DIR / d
        count, csv_path, xlsx_path = organize_folder(folder, rewrite_master=True)
        print(f"\n{d}: organized {count} leads")
        print(f"  CSV  -> {csv_path}")
        if xlsx_path.exists():
            print(f"  Excel-> {xlsx_path}")
        print(f"  Master-> {folder / 'all_results.csv'} (cleaned if not locked)")


def main() -> None:
    date_arg = sys.argv[1] if len(sys.argv) > 1 else None
    print_organize_report(date_arg)


if __name__ == "__main__":
    main()
